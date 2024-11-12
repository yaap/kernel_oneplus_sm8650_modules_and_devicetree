import os
import subprocess
import time
import shutil
import fnmatch
import pathlib
import sys
import csv
import codecs
import io
import glob
from datetime import datetime
import openpyxl
from openpyxl import load_workbook, Workbook

out_put = "out/oplus"
log_dir= 'LOGDIR/abi/'
exclude_file = ['*.ko', 'vmlinux']

def set_env_from_file(filepath):
    with open(filepath, 'r') as file:
        for line in file:
            # Remove any whitespace and newline characters, then check if the line contains environment variable definitions
            line = line.strip()
            if line and not line.startswith('#'):  # Ignore blank lines and comments
                key, _, value = line.partition('=')
                if key and value:
                    # Set the environment variable in the current process's environment
                    os.environ[key] = value
                    # Print the variable (if required)
                    #print("Set {}={}".format(key, value))

def delete_all_exists_file(folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    for item in os.listdir(folder_path):
        item_path = os.path.join(folder_path, item)
        if os.path.isdir(item_path):
            shutil.rmtree(item_path)
        else:
            os.remove(item_path)

def create_symlinks_from_pairs(src_and_targets):
    for src_pattern, target_filename in src_and_targets:
        src_files = glob.glob(src_pattern)
        directory = os.path.dirname(target_filename)
        if not os.path.exists(directory):
           os.makedirs(directory)

        if len(src_files) == 1:
            src_filename = src_files[0]

            if not os.path.exists(src_filename):
                print("Source file does not exist: {}".format(src_filename))
                continue

            src_absolute_path = pathlib.Path(src_filename).resolve()  # obtain absolute path
            if os.path.exists(target_filename):
                os.remove(target_filename)
            os.symlink(str(src_absolute_path), target_filename)
        elif len(src_files) > 1:
            print("Multiple source files match the pattern '{}'. Cannot create symlink.".format(src_pattern))
        else:
            print("No source files match the pattern '{}'. Cannot create symlink.".format(src_pattern))

def link_files_from_directories(src_dir_list, file_pattern, target_dir):
    target_path = pathlib.Path(target_dir)
    target_path.mkdir(parents=True, exist_ok=True)
    for src_dir in src_dir_list:
        src_path = pathlib.Path(src_dir)
        for file in src_path.rglob(file_pattern):
            file_absolute_path = file.resolve()  # obtain absolute path
            symlink_target = target_path / file.name
            if symlink_target.exists():
                os.remove(str(symlink_target))
            os.symlink(str(file_absolute_path), str(symlink_target))

def generate_abi(output_dir):
    generate_abi_gki_aarch64_oplus = "kernel_platform/build/abi/extract_symbols {}/ "\
                                             "--skip-module-grouping --symbol-list "\
                                             "{}/abi_gki_aarch64_oplus "\
                                             "| grep 'Symbol '>> "\
                                             "{}/abi_required_full.txt".format(output_dir, output_dir, output_dir)
    process = subprocess.Popen(generate_abi_gki_aarch64_oplus, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate()
    #print("stand output:{}".format(stdout.decode('utf-8')))
    #print("stand error:{}".format(stderr.decode('utf-8')))

def get_symbols_and_ko_list(input_file, output_list, output_ko, output_symbol_ko, delimiter=' '):
    unique_symbols = set()
    unique_ko = set()
    unique_symbols_ko = set()
    with open(input_file, 'r') as f_in:
        for line in f_in:
            columns = line.strip().split(delimiter)
            if len(columns) >= 4:
                unique_symbols.add(columns[1])
                unique_ko.add(columns[4])
                entry = (columns[1], columns[4])
                unique_symbols_ko.add(entry)

    with open(output_list, 'w') as f_out:
        #print("missing symbols is:")
        for item in unique_symbols:
            #print(item)
            f_out.write(item + '\n')

    with open(output_ko, 'w') as ko_out:
        #print("symbols is needed by:")
        for item in unique_ko:
            #print(item)
            ko_out.write(item + '\n')

    with open(output_symbol_ko, 'w') as s_ko_out:
        for symbol, ko in unique_symbols_ko:
            s_ko_out.write(symbol +' ' + ko + "\n")


def format_duration(td):
    minutes, seconds = divmod(td.seconds, 60)
    hours, minutes = divmod(minutes, 60)
    return '{:02} Hour {:02} Minute {:02} Second'.format(hours, minutes, seconds)


def compare_and_link_kos(kernel_dir, base_file, output_dir):
    # 以 Path 对象制造完整路径
    kernel_dir_path = pathlib.Path(kernel_dir)
    output_dir_path = pathlib.Path(output_dir)

    # 确保输出目录存在
    output_dir_path.mkdir(parents=True, exist_ok=True)

    # 获取所有的 .ko 文件
    ko_files = list(kernel_dir_path.glob('**/*.ko'))

    # 将检索到的文件输出到文件列表中
    list_file = output_dir_path / "current_ko_list.txt"
    list_file.write_text("\n".join(str(file) for file in ko_files))

    # 读取基准文件名
    with open(base_file, 'r') as f:
        base_kos = set(f.read().splitlines())

    # 对比文件并输出新增到文件中
    new_ko_file = output_dir_path / "new_ko_list.txt"
    with new_ko_file.open('w') as f:
        for ko_file in ko_files:
            if ko_file.name not in base_kos:
                f.write(ko_file.name + '\n')
                # 创建软链接
                link_path = output_dir_path / ko_file.name
                if not link_path.exists():
                    link_path.symlink_to(ko_file.resolve())

def extract_first_column(input_file, output_file):
    """
    读取 CSV 文件的每一行，提取每行的第一列，并将其输出到另一个文件中。

    :param input_file: 输入 CSV 文件的路径
    :param output_file: 输出文件的路径，将包含提取的第一列数据
    """
    try:
        # 使用 'with' 语句确保文件正确打开和关闭
        with open(input_file, 'r', encoding='utf-8') as source_file, open(output_file, 'w', encoding='utf-8') as target_file:
            for line in source_file:
                # 使用 strip() 移除尾部的换行符，然后以逗号分割每一行以获取列数据
                columns = line.strip().split(',')
                if columns:  # 确保不处理空行
                    first_column = columns[0]  # 提取第一列
                    target_file.write(first_column + '\n')  # 将第一列数据写入输出文件，并添加换行符
    except IOError as e:
        print("file error: {}".format(e))



def find_common_lines_and_write_sorted(file1_path, file2_path, output_file_path):
    # 使用 Path 对象来操作文件路径
    file1 = pathlib.Path(file1_path)
    file2 = pathlib.Path(file2_path)
    output_file = pathlib.Path(output_file_path)

    # 读取文件内容到集合中
    with open(file1, 'r') as f:
        lines_file1 = set(f.read().splitlines())
    with open(file2, 'r') as f:
        lines_file2 = set(f.read().splitlines())

    # 找到两个文件共有的行
    common_lines = lines_file1.intersection(lines_file2)

    # 对结果进行排序
    sorted_common_lines = sorted(common_lines)

    # 将排序后的共有行写入输出文件
    with open(output_file, 'w') as f:
        f.writelines("\n".join(sorted_common_lines))

def compare_files(file_a, file_b, common_file, only_a_file, only_b_file):
    """
    对比两个文件并输出结果到三个不同的文件中。

    :param file_a: 文件 A 的路径
    :param file_b: 文件 B 的路径
    :param common_file: 公共行将被写入的文件路径
    :param only_a_file: 只存在于文件 A 的行将被写入的文件路径
    :param only_b_file: 只存在于文件 B 的行将被写入的文件路径
    """
    with open(file_a, 'r', encoding='utf-8') as a, open(file_b, 'r', encoding='utf-8') as b:
        # 读取两个文件的内容到集合中，自动去除重复的行
        lines_a = set(a.readlines())
        lines_b = set(b.readlines())

    # 找到共同的行和独有的行
    common_lines = lines_a.intersection(lines_b)
    only_a_lines = lines_a - lines_b
    only_b_lines = lines_b - lines_a

    # 将结果写入到三个不同的文件，并进行排序
    with open(common_file, 'w', encoding='utf-8') as cf:
        cf.writelines(sorted(common_lines))

    with open(only_a_file, 'w', encoding='utf-8') as oaf:
        oaf.writelines(sorted(only_a_lines))

    with open(only_b_file, 'w', encoding='utf-8') as obf:
        obf.writelines(sorted(only_b_lines))

def txt_to_excel(txt_file, excel_file):
    """
    将一个文本文件转换为Excel文件，文本内容以空格分割。

    :param txt_file: 输入的文本文件路径
    :param excel_file: 输出的Excel文件路径
    """
    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    with open(txt_file, 'r', encoding='utf-8') as file:
        for row_index, line in enumerate(file, start=1):
            # 使用空格对每一行进行分割
            columns = line.strip().split(" ")
            for col_index, cell_value in enumerate(columns, start=1):
                # 将分割后的数据写入Excel工作表
                ws.cell(row=row_index, column=col_index).value = cell_value

    # 保存Excel工作簿
    wb.save(excel_file)
    #print("{} has been converted to {}".format(txt_file,excel_file))

def compare_and_update_excel(common_file, ko_owner_file):
    # 打开 common.xlsx 文件
    common_wb = load_workbook(common_file)
    common_ws = common_wb.active

    # 打开 ko_owner_list.xlsx 文件
    ko_owner_wb = load_workbook(ko_owner_file)
    ko_owner_ws = ko_owner_wb.active

    # 对 common.xlsx 的每一行进行检查和更新
    for row in common_ws.iter_rows(min_row=1, min_col=2, max_col=2):
        common_cell = row[0]
        # 在 ko_owner_list.xlsx 中查找对应的行
        for ko_row in ko_owner_ws.iter_rows(min_row=1):
            ko_owner_cell = ko_row[0]
            if ko_owner_cell.value == common_cell.value:
                # 如果找到匹配，则更新 common.xlsx 的相应单元格
                common_ws.cell(row=common_cell.row, column=3, value=ko_row[1].value)
                common_ws.cell(row=common_cell.row, column=4, value=ko_row[2].value)
                common_ws.cell(row=common_cell.row, column=5, value=ko_row[3].value)
                break

    # 保存更新后的 common.xlsx 文件
    #common_wb.save('updated_common.xlsx')
    common_wb.save(common_file)

def merge_excel_sheets(input_files, output_file):
    # 创建一个新的工作簿
    output_wb = Workbook()

    # 移除默认创建的空sheet
    output_wb.remove(output_wb.active)

    for file_path in input_files:
        # 加载每个Excel文件
        input_wb = load_workbook(file_path)
        # 获取第一个工作表
        input_ws = input_wb.active
        # 创建一个新的工作表，名称与输入文件相同
        sheet_name = file_path.split('/')[-1].replace('.xlsx', '')
        output_ws = output_wb.create_sheet(title=sheet_name)

        # 复制单元格数据到新的工作表
        for row in input_ws.iter_rows():
            for cell in row:
                output_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)

    # 保存这个新的Excel文件
    output_wb.save(output_file)


def main():
    start_time = datetime.now()
    print("Begin time:", start_time.strftime("%Y-%m-%d %H:%M:%S"))
    print("get symbol list start ...")

    os.system("./kernel_platform/oplus/tools/oplus_get_ko_list.sh  pineapple user")

    set_env_from_file('kernel_platform/common/build.config.constants')
    print("clang version:" + os.environ.get('CLANG_VERSION'))
    CLANG_PREBUILT_BIN = "kernel_platform/prebuilts/clang/host/linux-x86/clang-" + os.environ.get('CLANG_VERSION') + "/bin/:"
    print("clang path is:" + CLANG_PREBUILT_BIN)
    os.environ["PATH"] = "kernel_platform/build/build-tools/path/linux-x86:" + os.environ["PATH"]
    os.environ["PATH"] = CLANG_PREBUILT_BIN + os.environ["PATH"]

    print("qcom platform\n")

    print("[1] delete_all_exists_file...")
    delete_all_exists_file(out_put)

    # example: Predefining source and target files
    src_and_targets = [
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux.symvers', 'out/oplus/qcom/all/vmlinux.symvers'),
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux', 'out/oplus/qcom/all/vmlinux'),
    ]
    create_symlinks_from_pairs(src_and_targets)
    print("[2] create_symlinks for vmlinx to qcom all...")

    src_dir_list = [
        pathlib.Path("kernel_platform/oplus/prebuild/out/qcom/"),
    ]
    target_dir = pathlib.Path(out_put)/"qcom"/"all"
    link_files_from_directories(src_dir_list, "*.ko", target_dir)
    print("[3] create symlinks for qcom all ko...")

    extern_dir="/qcom/all"
    print("[4] generate qcom all abi...")
    generate_abi(out_put+extern_dir)

    print("[5] get qcom all symbols and ko list...")
    file_path = "{}/abi_required_full.txt".format(out_put+extern_dir)
    output_list = "{}/abi_required_symbol.txt".format(out_put+extern_dir)
    output_ko = "{}/abi_required_ko.txt".format(out_put+extern_dir)
    output_symbol_ko = "{}/abi_required_symbol_ko.txt".format(out_put+extern_dir)
    get_symbols_and_ko_list(file_path, output_list, output_ko, output_symbol_ko)

    src_and_targets = [
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux.symvers', 'out/oplus/qcom/oplus/vmlinux.symvers'),
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux', 'out/oplus/qcom/oplus/vmlinux'),
    ]
    print("[6] create_symlinks for qom oplus vmlinx ...")
    create_symlinks_from_pairs(src_and_targets)

    compare_and_link_kos(
        kernel_dir='kernel_platform/oplus/prebuild/out/qcom/',
        base_file='kernel_platform/oplus/tools/qcom_base_ko.txt',
        output_dir='out/oplus/qcom/oplus'
    )

    extern_dir="/qcom/oplus"
    print("[7] generate qcom oplus abi...")
    generate_abi(out_put+extern_dir)

    print("[8] get qcom oplus symbols and ko list...")
    file_path = "{}/abi_required_full.txt".format(out_put+extern_dir)
    output_list = "{}/abi_required_symbol.txt".format(out_put+extern_dir)
    output_ko = "{}/abi_required_ko.txt".format(out_put+extern_dir)
    output_symbol_ko = "{}/abi_required_symbol_ko.txt".format(out_put+extern_dir)
    get_symbols_and_ko_list(file_path, output_list, output_ko, output_symbol_ko)

    print("[9] get qcom platform oplus require symbols...")
    find_common_lines_and_write_sorted(
        'out/oplus/qcom/oplus/abi_required_symbol_ko.txt',
        'out/oplus/qcom/all/abi_required_symbol_ko.txt',
        'out/oplus/qcom/output_sorted_common_lines.txt'
    )

    # mtk plaform
    print("\nmtk platform\n")
    print("[1] null")
    # example: Predefining source and target files
    src_and_targets = [
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux.symvers', 'out/oplus/mtk/all/vmlinux.symvers'),
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux', 'out/oplus/mtk/all/vmlinux'),
    ]
    create_symlinks_from_pairs(src_and_targets)
    print("[2] create_symlinks for vmlinx to mtk all...")

    src_dir_list = [
        pathlib.Path("kernel_platform/oplus/prebuild/out/mtk/"),
    ]
    target_dir = pathlib.Path(out_put)/"mtk"/"all"
    link_files_from_directories(src_dir_list, "*.ko", target_dir)
    print("[3] create symlinks for mtk all ko...")

    extern_dir="/mtk/all"
    print("[4] generate mtk all abi...")
    generate_abi(out_put+extern_dir)

    print("[5] get mtk all symbols and ko list...")
    file_path = "{}/abi_required_full.txt".format(out_put+extern_dir)
    output_list = "{}/abi_required_symbol.txt".format(out_put+extern_dir)
    output_ko = "{}/abi_required_ko.txt".format(out_put+extern_dir)
    output_symbol_ko = "{}/abi_required_symbol_ko.txt".format(out_put+extern_dir)
    get_symbols_and_ko_list(file_path, output_list, output_ko, output_symbol_ko)

    src_and_targets = [
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux.symvers', 'out/oplus/mtk/oplus/vmlinux.symvers'),
        ('kernel_platform/oplus/platform/aosp_gki/vmlinux', 'out/oplus/mtk/oplus/vmlinux'),
    ]
    print("[6] create_symlinks for mtk oplus vmlinx ...")
    create_symlinks_from_pairs(src_and_targets)

    extract_first_column("kernel_platform/oplus/tools/ko_order_table.csv",
                         "out/oplus/mtk/oplus/mtk_base_ko.txt")
    compare_and_link_kos(
        kernel_dir='kernel_platform/oplus/prebuild/out/mtk/',
        base_file='out/oplus/mtk/oplus/mtk_base_ko.txt',
        output_dir='out/oplus/mtk/oplus'
    )

    extern_dir="/mtk/oplus"
    print("[7] generate mtk oplus abi...")
    generate_abi(out_put+extern_dir)

    print("[8] get mtk oplus symbols and ko list...")
    file_path = "{}/abi_required_full.txt".format(out_put+extern_dir)
    output_list = "{}/abi_required_symbol.txt".format(out_put+extern_dir)
    output_ko = "{}/abi_required_ko.txt".format(out_put+extern_dir)
    output_symbol_ko = "{}/abi_required_symbol_ko.txt".format(out_put+extern_dir)
    get_symbols_and_ko_list(file_path, output_list, output_ko, output_symbol_ko)

    print("[9] get mtk platform oplus require symbols...")
    find_common_lines_and_write_sorted(
        'out/oplus/mtk/oplus/abi_required_symbol_ko.txt',
        'out/oplus/mtk/all/abi_required_symbol_ko.txt',
        'out/oplus/mtk/output_sorted_common_lines.txt'
    )

    print("get all need symbol symbols...")
    compare_files('out/oplus/qcom/output_sorted_common_lines.txt',
                  'out/oplus/mtk/output_sorted_common_lines.txt',
                  'out/oplus/common.txt',
                  'out/oplus/qcom_only.txt',
                  'out/oplus/mtk_only.txt')

    txt_to_excel('out/oplus/common.txt', 'out/oplus/common.xlsx')
    txt_to_excel('out/oplus/qcom_only.txt', 'out/oplus/qcom_only.xlsx')
    txt_to_excel('out/oplus/mtk_only.txt', 'out/oplus/mtk_only.xlsx')

    compare_and_update_excel('out/oplus/common.xlsx', 'kernel_platform/oplus/tools/ko_owner_list.xlsx')
    compare_and_update_excel('out/oplus/qcom_only.xlsx', 'kernel_platform/oplus/tools/ko_owner_list.xlsx')
    compare_and_update_excel('out/oplus/mtk_only.xlsx', 'kernel_platform/oplus/tools/ko_owner_list.xlsx')
    input_files = [
        'out/oplus/common.xlsx',
        'out/oplus/qcom_only.xlsx',
        'out/oplus/mtk_only.xlsx'
    ]
    output_file = 'out/oplus/android15-6.6.xlsx'
    merge_excel_sheets(input_files, output_file)

    print("get symbol list end ...")
    end_time = datetime.now()
    print("End time:", end_time.strftime("%Y-%m-%d %H:%M:%S"))
    elapsed_time = end_time - start_time
    print("Total time:", format_duration(elapsed_time))

if __name__ == "__main__":
    main()
